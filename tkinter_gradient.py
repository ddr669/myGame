import numpy as np
import tkinter as tk    # py3
from ctypes import windll, byref, sizeof, c_int
import ctypes as ct
from customtkinter import *
from os import environ as AMBIENTE
import socket
import pandas as pd
from PIL import Image        
from os import system
import openpyxl
from time import localtime
#+++++++++++++++++++++++++++++++++++++++++++++++#
class Example(tk.Frame):
    def __init__(self, parent):
        tk.Frame.__init__(self, parent)
        #f1 = GradientFrame(self, borderwidth=1, relief="sunken")
        f2 = GradientFrame(self,(161, 183, 237),(175, 237, 161), borderwidth=1, relief="sunken")
        #f1.pack(side="top", fill="both", expand=True)
        f2.pack(side="bottom", fill="both", expand=True)

class GradientFrame(tk.Canvas):
    '''A gradient frame which uses a canvas to draw the background'''
    # recebe um canva(TKinter) e colore ele :3
    def __init__(self, parent, color1=(175, 237, 161), color2=(161, 183, 237),axi=1, **kwargs):
        tk.Canvas.__init__(self, parent, **kwargs)
        self.axi = axi
        # converte inteiro para hexadecimal
        def convHEX(r,g,b):
            return f"#{r:02x}{g:02x}{b:02x}"

        
        if type(color1[0]) is str:
            self._color1 = color1
            self._color2 = color2
        else:
            self._color1 = convHEX(int(color1[0]),int(color1[1]),int(color1[2]))
            self._color2 = convHEX(int(color2[0]),int(color2[1]),int(color2[2]))
            
        self.bind("<Configure>", self._draw_gradient)


    def _draw_gradient(self, event=None):
        '''Draw the gradient'''
        self.delete("gradient")
        width = self.winfo_width()
        height = self.winfo_height()
        
        if self.axi == 1:
            limit = height
            (r1,g1,b1) = self.winfo_rgb(self._color1)
            (r2,g2,b2) = self.winfo_rgb(self._color2)
            r_ratio = float(r2-r1) / limit
            g_ratio = float(g2-g1) / limit
            b_ratio = float(b2-b1) / limit
            for i in range(limit):
                
                nr = int(r1 + (r_ratio * i))
                ng = int(g1 + (g_ratio * i))
                nb = int(b1 + (b_ratio * i))
                color = "#%4.4x%4.4x%4.4x" % (nr,ng,nb)
                self.create_line(0,i,i+width*height,height, tags=("gradient",), fill=color)
            self.lower("gradient")
        else:
            limit = width
            (r1,g1,b1) = self.winfo_rgb(self._color1)
            (r2,g2,b2) = self.winfo_rgb(self._color2)
            r_ratio = float(r2-r1) / limit
            g_ratio = float(g2-g1) / limit
            b_ratio = float(b2-b1) / limit
            for i in range(limit):
                nr = int(r1 + (r_ratio * i))
                ng = int(g1 + (g_ratio * i))
                nb = int(b1 + (b_ratio * i))
                color = "#%4.4x%4.4x%4.4x" % (nr,ng,nb)
                self.create_line(i,0,i,height, tags=("gradient",), fill=color)
            self.lower("gradient")
    def convHEX(r,g,b):
        return f"#{r:02x}{g:02x}{b:02x}"
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++#       
if __name__ == "__main__":
    class App:
        def __init__(self):
            
            
            self.qnt_value = 0
            self.plusplus = CTkImage(Image.open("plus.png"))
            self.lessless = CTkImage(Image.open("less.png"))
           
            
            self.vendidos = AMBIENTE["HOMEPATH"]+"\\plan\\vendidos.xlsx"
            self.balcao   = AMBIENTE["HOMEPATH"]+"\\plan\\balcao.xlsx"
            self.estoque  = AMBIENTE["HOMEPATH"]+"\\plan\\estoque.xlsx"
            self.root = CTk()
            self.root.geometry("400x400")
            self.root.title("Lançador de comandas")
            self.root.resizable(False, False)  

            self.label = tk.Label(master=self.root,text="oI", width=400, height=400,relief=None)
            self.label.pack(fill="both", expand=True)#relx=0.5, rely=0.5, anchor="center")
    
            self.dark_title_bar(self.root)
            self.convert = lambda r,g,b: f"#{r:02x}{g:02x}{b:02x}"

    
            Example(self.label).pack(fill="both", expand=True)
    
            self.tela = tk.Label(master=self.root, width=40,height=20
                       , text="", background=self.convert(240,230,230))
            self.tela.place(relx=0.5, rely=0.5, anchor="center")
        def log(self, acao):
            with open(AMBIENTE["HOMEPATH"]+"\\log\\log.txt", "a") as p:
                p.write("[{}/{}/{}:=:{}:{}]: {}\n".format(localtime()[0],
                                                      localtime()[1],
                                                      localtime()[2],
                                                      localtime()[3],
                                                      localtime()[4],
                                                        acao))
        def deposito(self):
            tela2 = tk.Label(master=self.root, width=50, height=20)
            tela2.place(relx=0.5, rely=0.5, anchor="center")
            
            
        def mais(self):
               
            self.qnt_value += 1
            self.quantidade.configure(text=self.qnt_value)

        def menos(self):
         
            self.qnt_value -= 1
            self.quantidade.configure(text=self.qnt_value)

            
        def get_json(self):
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.bind("localhost", 1111)
            s.listen(5)
            client, address = s.accept()
        
            return json

        def dark_title_bar(self, window):
            window.update()
            DWMWA_USE_IMMERSIVE_DARK_MODE = 20
            set_window_attribute = ct.windll.dwmapi.DwmSetWindowAttribute
            get_parent = ct.windll.user32.GetParent
            hwnd = get_parent(window.winfo_id())
            rendering_policy = DWMWA_USE_IMMERSIVE_DARK_MODE
            value = 2
            value = ct.c_int(value)
            set_window_attribute(hwnd, rendering_policy, ct.byref(value), ct.sizeof(value))



        def main(self):#arqvJson):
            
            self.lista_obj = CTkComboBox(self.tela, values=["escolha um produto"])
            self.lista_obj.place(relx=0.3, rely=0.1, anchor="center")
            self.ler()
            self.bottao_lan = CTkButton(self.tela, text="Lançar", width=30, command=self.escrever)
            self.bottao_lan.place(relx=0.3, rely=0.8, anchor="center")

            self.bottao_op = CTkButton(self.tela, text="Depósito", width=55, fg_color=self.convert(200,80,60), command=self.deposito)
            self.bottao_op.place(relx=0.8, rely=0.8, anchor="center")

            self.bottao_pls = CTkButton(self.tela, image=self.plusplus,text="", width=20, height=20, corner_radius=20, border_width=1, command=self.mais)
            self.bottao_pls.place(relx=0.5, rely=0.3, anchor="center")

            self.bottao_lss = CTkButton(self.tela, image=self.lessless, text="", width=20, height=20, corner_radius=20, border_width=1, command=self.menos)
            self.bottao_lss.place(relx=0.1, rely=0.3, anchor="center")

            self.quantidade = CTkLabel(self.tela, text=self.qnt_value)
            self.quantidade.place(relx=0.3, rely=0.3, anchor="center")
    
         
                    
            self.root.mainloop()

        def ler(self, oque=0,arqv=AMBIENTE["HOMEPATH"]+"\\plan\\balcao.xlsx"):
            if oque == 0: # balcao
                try:
                    excel = pd.read_excel(arqv)
                    self.dic   = excel.to_dict()["Produtos"]
                    tesao = ""
    
                    self.contagem = excel.to_dict()["Quantidade"]
                    self.prod     = excel.to_dict()["Produtos"]
                    dick = []
                    for a in range(0, len(self.prod)):
                        dick.append(self.prod[a])

            
                    self.lista_obj.configure(values=dick)
        
                    self.preco    = excel.to_dict()["Preço"]
                except:
                    print("error, provavelmente sem arquivo")
                    self.ler(1)
                    
                for a in range(0,len(self.dic)):
                
                
                    if str(self.contagem[a]) == "nan":
                        pass
                    else:
                        print(self.contagem[a])            

            
          
            elif oque == 1: # vendidos
                try:
                    tfile = openpyxl.load_workbook(AMBIENTE["HOMEPATH"]+"\\plan\\vendidos.xlsx")
                    current = tfile["new_sheet"]
                    data = int(localtime()[2])
                    self.contagem = excel.to_dict()[localtime()[2]]
                    self.prod     = excel.to_dict()["Produtos"]
                except:
                    print("ish...")
                    print(current)
                    def _find2():
                        for column in "ABCDEFGHIJKLMNOPQRSTU":
                            if current[column+"1"].value == data:
                                return column

                            elif current[column+"1"].value == None:
                                current[column+"1"] = data
                                self.log("data inserida na planinha")
                                return column

                    def _find():
                        for row in range(1, current.max_row+1):
                            for column in "ABC":
                                cell = "{}{}".format(column, row)
                                if current[cell].value == self.prod:
                                    return cell
                                
                    #coluna = _find()[1:]
           
                    c = _find2()
                    print(c)
                    self.log("arquivo vendidos foi salvo")
                    tfile.save(AMBIENTE["HOMEPATH"]+"\\plan\\vendidos.xlsx")
                    
                #arqv=AMBIENTE["HOMEPATH"]+"\\plan\\vendidos.xlsx"
                #self.prod = pd.read_excel(arqv)

                return current
            
        def escrever(self):
            vendidos = self.ler(oque=1)
        
            if localtime()[2] in vendidos.columns:
                data = int(localtime()[2])
                qnt  = str(self.quantidade.cget("text")) # qnt = columns
                prod = str(self.lista_obj.get()) # prod = index
                print("="*20+"\n",prod, qnt, data+"\n"+"="*20)
                bfile = openpyxl.load_workbook(AMBIENTE["HOMEPATH"]+"\\plan\\balcao.xlsx")
                tfile = openpyxl.load_workbook(AMBIENTE["HOMEPATH"]+"\\plan\\vendidos.xlsx")
                current = tfile["new_sheet"]
                def _find():
                    for row in range(1, current.max_row+1):
                        for column in "ABC":
                            cell = "{}{}".format(column, row)
                            if current[cell].value == prod:
                                return cell
            
                coluna = _find()[1:]
                print(coluna)
                def _find2():
                    for column in "ABCDEFGHIJKLMNOPQRSTU":
                        if current[column+"1"].value == data:
                            return column

                        elif current[column+"1"].value == None:
                            current[column+"1"] = data
                            return column
             

                c = _find2()
                print(c)
            
                q = current[str(c)+str(coluna)].value
                if q:
                    current[str(c)+str(coluna)] = int(q) + int(qnt)
                    self.log("foi vendido {} de {}".format(qnt, current["A"+str(coluna)].value)) 
                else:
                   
                    current[str(c)+str(coluna)] = int(qnt)
                    self.log("vendidos {} do prod {}".format(qnt, current["A"+str(coluna)].value))

                self.log("arquivo vendidos salvo")
                tfile.save(AMBIENTE["HOMEPATH"]+"\\plan\\vendidos.xlsx")
            
                current = bfile["new_sheet"]
            
                coluna = _find()[1:]
                var_cell = "B"+str(coluna)
            
                res = int(current[var_cell].value) - int(qnt)
                self.log("vendidos {} de {} ".format(qnt, current["A"+str(coluna)].value))
                current[var_cell] = res
                bfile.save(AMBIENTE["HOMEPATH"]+"\\plan\\balcao.xlsx")
           
            
            else:
                
                data = int(localtime()[2])
                qnt  = str(self.quantidade.cget("text")) # qnt = columns
                print(qnt)
                prod = str(self.lista_obj.get()) # prod = index
            
                bfile = openpyxl.load_workbook(AMBIENTE["HOMEPATH"]+"\\plan\\balcao.xlsx")
                tfile = openpyxl.load_workbook(AMBIENTE["HOMEPATH"]+"\\plan\\vendidos.xlsx")
                current = tfile["new_sheet"]
            
                def _find():
                    for row in range(1, current.max_row+1):
                            cell = "{}{}".format("A", row)
                            if current[cell].value == prod:
                                return cell
                
                coluna = _find()[1:]
                print(coluna)
                def _find2():
                    for column in "ABCDEFGHIJKLMNOPQRSTU":
                        if current[column+"1"].value == data:
                            return column

                        elif current[column+"1"].value == None:
                            current[column+"1"] = data
                            return column
           
                c = _find2()
                print(c)
                q = current[str(c)+str(coluna)].value
                if q:
                    current[str(c)+str(coluna)] = int(q) + int(qnt)
                    self.log("vendido {} de {}".format(qnt, current["A"+str(coluna)].value))
                    
                else:
                    current[str(c)+str(coluna)] = int(qnt)
                    self.log("vendido {} de {}".format(qnt, current["A"+str(coluna)].value))
                print(type(q))
                def _find3():
                    for row in range(1, current.max_row+1):
                        cell = "A{}".format( row)
                        if current[cell].value == "Quantidade":
                            return cell
                c = _find3()
                print("=",c,"=")
                #current[str(c)+str(coluna)] = int(q) + int(qnt)
                tfile.save(AMBIENTE["HOMEPATH"]+"\\plan\\vendidos.xlsx")
                current = bfile["new_sheet"]
                coluna = _find()[1:]
        
                
                current["B"+str(coluna)] = float(current["B"+str(coluna)].value) - int(qnt)
                bfile.save(AMBIENTE["HOMEPATH"]+"\\plan\\balcao.xlsx")
            
            
            
        

            self.main()

                

a = App()
a.main()
